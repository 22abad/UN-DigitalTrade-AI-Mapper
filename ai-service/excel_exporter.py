"""Excel exporter — exports RDTII indicator mappings to Excel format.

Follows the RDTII 2.1 data collection practice format requirements:
  - Act and/or practice: English translation, official language (optional), year, number
  - Coverage: horizontal / sectoral
  - Impacts or comments: section, article, clause numbers
  - Timeframe: Month and year (came into force), amendments
  - Reference: official URLs

Supports:
  - Single-document export (list of IndicatorMapping → .xlsx)
  - Batch export (multiple documents → single .xlsx with multiple sheets)
  - Color-coded: new additions in blue, existing in black
"""

from __future__ import annotations

import io
import json
from typing import Any

from schemas import IndicatorMapping

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# RDTII 2.1 format columns
_RDTII_COLUMNS = [
    ("Pillar", 10),
    ("Indicator", 12),
    ("Score", 8),
    ("Act and/or practice", 45),
    ("Coverage", 14),
    ("Impacts or comments", 55),
    ("Timeframe", 30),
    ("Reference (URL)", 50),
    ("Verbatim Quote", 55),
    ("Provision Article", 18),
    ("Source Validation", 20),
]

# Column widths
_COL_WIDTHS = {name: width for name, width in _RDTII_COLUMNS}

# Styling
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_ALTERNATE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_NEW_ADDITION_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
_NEW_ADDITION_FONT = Font(name="Calibri", size=10, color="1B5E20")


def _init_workbook() -> Any:
    """Create a new workbook with header styling."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RDTII Mappings"
    
    header_names = [c[0] for c in _RDTII_COLUMNS]
    for col_idx, name in enumerate(header_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(name, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header_names))}1"
    
    return wb


def _source_legislation_display(mapping: IndicatorMapping) -> str:
    """Format the 'Act and/or practice' column per RDTII format."""
    title = mapping.source_legislation or ""
    return title


def _impact_comment(mapping: IndicatorMapping) -> str:
    """Format the 'Impacts or comments' column with provision + interpretation."""
    parts = []
    prefix = f"[{mapping.article_clause}] " if mapping.article_clause else ""
    if mapping.impact:
        # Prepend prefix directly to the first line (Interpretation)
        parts.append(f"{prefix}Interpretation: {mapping.impact}")
    else:
        parts.append(f"{prefix}Interpretation: Mapping for indicator {mapping.indicator}")
        
    if mapping.verbatim_quote:
        parts.append(f"Provision: {mapping.verbatim_quote[:200]}")
    return "\n".join(parts)


def _timeframe_display(mapping: IndicatorMapping) -> str:
    """Extract/format the Timeframe from mapping metadata."""
    features = mapping.features or {}
    if features.get("_timeframe_column"):
        return features.get("_timeframe_column")

    ts = mapping.timestamp_verification or {}
    if ts.get("verified"):
        best = ts.get("best_date", "")
        if best:
            from timeframe_extractor import format_to_month_year
            best_my = format_to_month_year(best)
            return f"Since {best_my}" if best_my else f"Since {best}"

    last_upd = mapping.last_update or ""
    if last_upd:
        from timeframe_extractor import format_to_month_year
        last_upd_my = format_to_month_year(last_upd)
        if last_upd_my:
            return f"Since {last_upd_my}"
    return "In force (date unknown)"


def _coverage_display(mapping: IndicatorMapping) -> str:
    """Format Coverage column."""
    return mapping.scope if mapping.scope != "unknown" else ""


def _source_validation_display(mapping: IndicatorMapping) -> str:
    """Show source validation grade from optional metadata."""
    features = mapping.features or {}
    src_valid = features.get("_source_grade", "")
    if src_valid:
        return f"Grade: {src_valid}"
    return ""


def _provision_article_display(mapping: IndicatorMapping) -> str:
    """Extract article/provision reference from mapping metadata."""
    if mapping.article_clause:
        return mapping.article_clause
    quote = mapping.verbatim_quote or ""
    return quote[:80] if quote else ""


def mappings_to_excel(
    mappings: list[IndicatorMapping],
    country: str = "",
    batch_id: str = "",
) -> io.BytesIO:
    """Export a list of IndicatorMapping objects to an Excel workbook.
    
    Returns BytesIO stream ready for download.
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = _init_workbook()
    ws = wb.active

    for row_idx, mapping in enumerate(mappings, 2):
        row_data = [
            f"Pillar {mapping.pillar}",
            mapping.indicator,
            mapping.score,
            _source_legislation_display(mapping),
            _coverage_display(mapping),
            _impact_comment(mapping),
            _timeframe_display(mapping),
            mapping.source_url,
            mapping.verbatim_quote,
            _provision_article_display(mapping),
            _source_validation_display(mapping),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER

            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = _ALTERNATE_FILL

        # Set row height for readability
        ws.row_dimensions[row_idx].height = 45

    # Auto-adjust some columns
    for col_idx, (name, _) in enumerate(_RDTII_COLUMNS, 1):
        if name in ("Impacts or comments", "Verbatim Quote"):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                _COL_WIDTHS.get(name, 20), 55
            )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def mappings_to_excel_with_sheets(
    country_data: dict[str, list[IndicatorMapping]],
) -> io.BytesIO:
    """Export mappings grouped by country into a multi-sheet workbook.
    
    Each country gets its own sheet.
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export.")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for country, mappings in country_data.items():
        ws = wb.create_sheet(title=country[:31])  # Excel sheet name max 31 chars
        header_names = [c[0] for c in _RDTII_COLUMNS]
        for col_idx, name in enumerate(header_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(name, 20)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header_names))}1"

        for row_idx, mapping in enumerate(mappings, 2):
            row_data = [
                f"Pillar {mapping.pillar}",
                mapping.indicator,
                mapping.score,
                _source_legislation_display(mapping),
                _coverage_display(mapping),
                _impact_comment(mapping),
                _timeframe_display(mapping),
                mapping.source_url,
                mapping.verbatim_quote,
                _provision_article_display(mapping),
                _source_validation_display(mapping),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = _THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = _ALTERNATE_FILL
            ws.row_dimensions[row_idx].height = 45

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def create_excel_response(
    mappings: list[IndicatorMapping],
    country: str = "unknown",
    filename: str = "rdtii_mappings.xlsx",
):
    """Create a FastAPI StreamingResponse for Excel download."""
    from fastapi.responses import StreamingResponse

    buf = mappings_to_excel(mappings, country=country)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )


__all__ = ["mappings_to_excel", "mappings_to_excel_with_sheets", "create_excel_response"]
